from django.shortcuts import render, redirect, get_object_or_404
from django.views import generic
from .models import Exercise, Workout, ExerciseType, Set
from .forms import WorkoutForm, ExerciseForm, SetForm, ExerciseSetForm
import ast
import json
from django.http import JsonResponse
from django.db import IntegrityError, transaction
from django.contrib import messages

# Create your views here.
class ExerciseList(generic.ListView):
    queryset = ExerciseType.objects.all()
    template_name = "excersise_list.html"
    paginate_by = 2

def workout_list(request):
    workouts = Workout.objects.all()
    return render(request, 'workouts/workout_list.html', {'workouts': workouts})


# Create a new workout
def create_workout(request):
    workout_title = request.GET.get('workout_title', '')  # Get workout_title from query parameters, default to an empty string if not provided
    if(workout_title):
        context = {
            'workout_title': workout_title,
        }
        return render(request, 'workouts/create_workout.html', context)
    return render(request, 'workouts/create_workout.html')



# Add exercises to a workout
def add_exercises(request, workout_title):
    print("workout title is:", workout_title)


    if request.method == 'POST':
        workout_title = request.POST.get('workout_title')  # Get workout title from the form data
    
    query = request.GET.get('q')
    if query:
        exercises = ExerciseType.objects.filter(exercise_name__icontains=query).exclude(exercise_image__isnull=True)
    else:
        exercises = ExerciseType.objects.exclude(exercise_image__isnull=True)
    
    context = {
        'exercises': exercises,
        'workout_title': workout_title,  # Replace with actual workout title
    }
    return render(request, 'workouts/add_exercises.html', context)


# Add sets to an exercise
def add_exercise_sets(request):
    if request.method == 'POST':
        workout_title = request.POST.get('workout_title')
        selected_excersises = request.POST.get('selected_excersises')
        
        # Convert the string of exercises into a list
        try:
            selected_excersises = ast.literal_eval(selected_excersises)
        except (ValueError, SyntaxError):
            selected_excersises = []  # Default to an empty list if conversion fails

        # Filter ExerciseType objects based on selected exercise names
        matching_exercises = ExerciseType.objects.filter(exercise_name__in=selected_excersises)

        
        context = {
            'selected_excersises': matching_exercises,
            'workout_title': workout_title,
        }
        return render(request, 'workouts/workout_set_reps.html', context)
    
    
def save_workout(request):
    if request.method == "POST":
        try:
            print("Data sent from front end: ", request.POST)
            

            # Start a transaction block
            with transaction.atomic():
                # First, validate and create Workout using the form
                workout_form = WorkoutForm(request.POST)
                if not workout_form.is_valid():
                    return JsonResponse({'status': 'error', 'message': 'Invalid workout data', 'errors': workout_form.errors})
                
                
                # Associate the workout with the logged-in user
                workout = workout_form.save(commit=False)
                workout.user = request.user  # Set the user field to the logged-in user
                workout.save()

                # Deserialize exercises JSON string into Python list
                exercises_json = request.POST.get("exercise_type")  # This is a string
                exercises = json.loads(exercises_json)

                # Process exercises and sets
                for exercise_data in exercises:
                    # Create Exercise using the form
                    exercise_form = ExerciseForm({
                        'workout': workout.id,
                        'exercise_type': ExerciseType.objects.get(exercise_name=exercise_data["exercise_type"]).id
                    })

                    if not exercise_form.is_valid():
                        # If any exercise form is invalid, roll back the entire transaction
                        raise IntegrityError("Invalid exercise data")
                    
                    # Save the exercise
                    exercise = exercise_form.save()

                    # Process sets
                    for set_data in exercise_data["set_reps"]:
                        set_form = SetForm({
                            'exercise': exercise.id,
                            'set_number': set_data["set"],
                            'reps': set_data["reps"]
                        })

                        if not set_form.is_valid():
                            # If any set form is invalid, roll back the entire transaction
                            raise IntegrityError("Invalid set data")
                        
                        # Save the set
                        set_form.save()
                        

                # If everything works, return success
                return JsonResponse({'status': 'success', 'message': 'Workout saved sucessfully.'})

        except ExerciseType.DoesNotExist as e:
            return JsonResponse({'status': 'error', 'message': 'Invalid exercise type.'})

        except IntegrityError as e:
            return JsonResponse({'status': 'error', 'message': str(e)})

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': 'An unexpected error occurred.'})
    
    # If the request method is not POST
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'})

def delete_workout(request, id):
    # Fetch the workout using the ID, and ensure the user owns the workout
    print("The view has been triggered")
    workout = get_object_or_404(Workout, id=id, user=request.user)
    print(f"delete workout view got the id of: {id}")

    print("workout to delete", workout)

    if workout.user == request.user:
        workout.delete()
        print("sucessfully deleted workout")
        
    return redirect('workouts')  # Redirect to the workouts page

def edit_workout(request, id):

    # Fetch the workout using the ID, and ensure the user owns the workout
    print("edit workout id is:", id)
    workout = get_object_or_404(Workout, id=id, user=request.user)
    print("edit workout id is:", workout.title)

    # Get the exercises in the workout
    exercises = Exercise.objects.filter(workout=workout)
    exercise_types = ExerciseType.objects.filter(id__in=exercises.values('exercise_type'))
    # Get the sets associated with each exercise
    exercises_with_sets = []
    for exercise in exercises:
        sets = Set.objects.filter(exercise=exercise)

        exercises_with_sets.append({
            'exercise_name': exercise.exercise_type.exercise_name,  # Return the name of the exercise
            'sets': sets
        })

    return render(request, 'workouts/edit_workout.html', {'workout': workout, 'exercises_with_sets': exercises_with_sets})
    
    # print("excersises with sets:", exercises_with_sets)
    
    # print("Exercises in edit workout are:", exercises_with_sets)


    return render(request, 'workouts/edit_workout.html', {'workout': workout, 'excersises': exercises})
    
def update_workout(request):
    if(request.method == "POST"):
            print("request method was:", request.method)
            print("Reached the update workout view")
            workout_id = request.POST.get("id")  # This is a string
            print("Update workout id", workout_id)

            # Fetch the existing workout and check permissions
            workout = get_object_or_404(Workout, id=workout_id, user=request.user)
            print("Workout to update", workout)
            
            with transaction.atomic():
                # Validate and update workout
                workout_form = WorkoutForm(request.POST, instance=workout)
                if not workout_form.is_valid():
                    return JsonResponse({'status': 'error', 'message': 'Invalid workout data', 'errors': workout_form.errors})
                
                workout_form.save()
                print("Successfully updated workout title")

                # Delete old exercises and their sets
                old_exercises = Exercise.objects.filter(workout=workout)
                for old_exercise in old_exercises:
                    Set.objects.filter(exercise=old_exercise).delete()
                old_exercises.delete()
                print("Successfully deleted old exercises and sets")

                # Deserialize exercises JSON string into Python list
                exercises_json = request.POST.get("exercise_type")  # This is a string
                exercises = json.loads(exercises_json)

                # Process exercises and sets
                for exercise_data in exercises:
                    # Create Exercise using the form
                    exercise_form = ExerciseForm({
                        'workout': workout.id,
                        'exercise_type': ExerciseType.objects.get(exercise_name=exercise_data["exercise_type"]).id
                    })

                    if not exercise_form.is_valid():
                        # If any exercise form is invalid, roll back the entire transaction
                        raise IntegrityError("Invalid exercise data")
                    
                    # Save the exercise
                    exercise = exercise_form.save()

                    # Process sets
                    for set_data in exercise_data["set_reps"]:
                        set_form = SetForm({
                            'exercise': exercise.id,
                            'set_number': set_data["set"],
                            'reps': set_data["reps"]
                        })

                        if not set_form.is_valid():
                            # If any set form is invalid, roll back the entire transaction
                            raise IntegrityError("Invalid set data")
                        
                        # Save the set
                        set_form.save()
            
    
    # If the request method is not POST
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'})


import openpyxl
from django.shortcuts import render
from .models import Exercise

# def import_excel_data(request):
#     if request.method == "POST" and request.FILES["file"]:
#         # Get the uploaded file
#         file = request.FILES["file"]
#         wb = openpyxl.load_workbook(file)
#         sheet = wb.active

#         for row in sheet.iter_rows(min_row=2, values_only=True):
#             exercise_name, description_url, exercise_image, exercise_image1, muscle_group_details, muscle_group, equipment_details, equipment,  rating , description = row

#             # Create the Exercise objects
#             Exercise.objects.create(
#                 exercise_name=exercise_name,
#                 description_url=description_url,
#                 exercise_image=exercise_image,
#                 exercise_image1=exercise_image1,
#                 muscle_group_details=muscle_group_details,
#                 muscle_group=muscle_group,
#                 equipment_details=equipment_details,
#                 equipment=equipment,
#                 rating=rating,
#                 description=description
#             )
#         return render(request, 'workouts/success.html')  # A success page after import
#     return render(request, 'workouts/import_form.html')  # A form page where you upload the file

def import_excel_data(request):
    if request.method == "POST" and request.FILES["file"]:
        # Get the uploaded file
        file = request.FILES["file"]
        wb = openpyxl.load_workbook(file)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            exercise_name, description_url, exercise_image, exercise_image1, muscle_group_details, muscle_group, equipment_details, equipment, rating, description = row

            # Check if the exercise already exists in the database
            if not ExerciseType.objects.filter(exercise_name=exercise_name).exists():
                # Create the Exercise objects if it does not exist
                ExerciseType.objects.create(
                    exercise_name=exercise_name,
                    description_url=description_url,
                    exercise_image=exercise_image,
                    exercise_image1=exercise_image1,
                    muscle_group_details=muscle_group_details,
                    muscle_group=muscle_group,
                    equipment_details=equipment_details,
                    equipment=equipment,
                    rating=rating,
                    description=description
                )
            else:
                # Log a message or handle the case when duplicate exercise is found
                print(f"Exercise '{exercise_name}' already exists and will be skipped.")

        return render(request, 'workouts/success.html')  # A success page after import
    return render(request, 'workouts/import_form.html')  # A form page where you upload the file




    
